import openpyxl
from openpyxl.utils import get_column_letter


def write_data(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    column_names = ['Университет', 'Сайт программ', 'Название программы', 'Описание стратегического проекта',
                    'Цель стратегического проекта', 'Задачи стратегического проекта', 'Ожидаемые результаты стратегических проектов']
    for col_num, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=col_num, value=column_name)
        column_width = len(column_name) + 2
        sheet.column_dimensions[get_column_letter(col_num)].width = column_width

    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=value)
            if len(str(value)) + 2 > sheet.column_dimensions[get_column_letter(col_num)].width:
                sheet.column_dimensions[get_column_letter(col_num)].width = len(str(value)) + 2

    workbook.save('result.xlsx')
