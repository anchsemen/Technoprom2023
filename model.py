import graphviz
import openpyxl
import numpy as np
from sklearn.metrics import precision_score, accuracy_score, recall_score
from sklearn.tree import DecisionTreeClassifier, export_graphviz
from sklearn.model_selection import GridSearchCV, train_test_split


def find_value(search_word1):
    return [item[2] for item in tfidf_data if item[0] == search_word1]


def load_data(file_name):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Sheet']
    data = [row for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return data


def preprocess_data(data):
    data = sorted(data, key=lambda x: x[2], reverse=True)
    part_data = [item for item in data if item[2] >= 0.33]
    words = np.unique([item[1] for item in part_data])
    return part_data, words


def create_table(project_names, words, part_data):
    table = [[''] + list(words)]
    for project in project_names:
        row = [project] + [next((item[2] for item in part_data if item[0] == project and item[1] == word), 0) for word in words]
        table.append(row)
    return table


def write_output(file_name, table):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_index, row_data in enumerate(table, start=1):
        for col_index, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_index, column=col_index, value=cell_value)
    workbook.save(file_name)


def train_decision_tree(X_train, y_train):
    param_grid = {'max_depth': [5, 10, 15, 20]}
    grid_search = GridSearchCV(DecisionTreeClassifier(), param_grid, cv=3)
    grid_search.fit(X_train, y_train)
    best_params = grid_search.best_params_
    best_score = grid_search.best_score_
    return best_params, best_score


def visualize_decision_tree(tree, words):
    dot_data = export_graphviz(tree, out_file=None,
                               feature_names=words, class_names=['Базовая часть', 'Спецчасть'],
                               filled=True, rounded=True, special_characters=True)

    graph = graphviz.Source(dot_data)
    graph.render("decision_tree")
    graph.view()


def evaluate_model(tree, X_test, y_test):
    y_pred = tree.predict(X_test)
    precision = precision_score(y_test, y_pred)
    accuracy = accuracy_score(y_test, y_pred)
    recall = recall_score(y_test, y_pred)
    return precision, accuracy, recall


project_data = load_data('result.xlsx')
tfidf_data = load_data('tfidf.xlsx')

part_data, words = preprocess_data(tfidf_data)

project_names = [row[2] for row in project_data[1:]]
table = create_table(project_names, words, part_data)

write_output('output.xlsx', table)

project_labels = [row[7] for row in project_data[1:]]
X = [[cell.value for cell in row[1:]] for row in table[1:]]

X_train, X_test, y_train, y_test = train_test_split(X, project_labels, test_size=0.2, random_state=42)
best_params, best_score = train_decision_tree(X_train, y_train)

print("Наилучшие параметры:", best_params)
print("Best score is:", best_score)

tree = DecisionTreeClassifier(max_depth=5).fit(X_train, y_train)

visualize_decision_tree(tree, words)

precision, accuracy, recall = evaluate_model(tree, X_test, y_test)
print(f"Precision: {precision}", f"Accuracy: {accuracy}", f"Recall: {recall}", sep='\n')
