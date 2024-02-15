import openpyxl
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer

workbook = openpyxl.load_workbook('text preprocessing.xlsx')
sheet = workbook['Sheet']

column_values = [row[0] for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)]

column_data, array_tfidf, arr = [], [], []
for row in sheet.iter_rows(min_row=2, values_only=True):
    row_data = ' '.join(str(cell) for cell in row[3:7] if cell is not None)
    column_data.append(row_data)

data = [row[0].value for row in sheet.iter_rows(min_row=2, min_col=8, max_col=8)]

vectorizer = TfidfVectorizer()
tfidf_matrix = vectorizer.fit_transform(column_data)
words = vectorizer.get_feature_names_out()
tfidf_array = tfidf_matrix.toarray()

temp_arr = [['Стратегический проект', 'Слово', 'Частота']]


for i, doc in enumerate(column_data):
    sorted_indices = tfidf_array[i].argsort()[::-1][:100]  # Limit to top 100
    for j in sorted_indices:
        tfidf_value = tfidf_array[i][j]
        if tfidf_value > 0:
            word = words[j]
            temp_arr.append([column_values[i], word, tfidf_value])
            array_tfidf.append(tfidf_value)
            arr.append(data[i])

workbook = openpyxl.Workbook()
sheet = workbook.active
for row_index, row_data in enumerate(temp_arr, start=1):
    for col_index, cell_value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=cell_value)

max_length = max(len(str(cell_value)) for row_data in temp_arr for cell_value in row_data)
sheet.column_dimensions[get_column_letter(2)].width = max_length

workbook.save('tfidf.xlsx')
workbook.close()
